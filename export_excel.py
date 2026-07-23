#!/usr/bin/env python3
"""
FlightPath Infogain — Daily Work Log Excel exporter
Dashboard + daily detail + month filters. Excludes ProductiveIT / personal.

Usage:
  python export_excel.py
  python export_excel.py --month 2026-07
  python export_excel.py --date 2026-07-20
  python export_excel.py --from 2026-04-08 --to 2026-07-20
"""

from __future__ import annotations

import argparse
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent
OUT_DIR = ROOT / "exports"

START = date(2026, 4, 8)
END = date(2026, 7, 23)

LEAVE = {date(2026, 5, 14), date(2026, 6, 20), date(2026, 7, 8)}
ON_MOVEMENT = {date(2026, 7, 9), date(2026, 7, 10)}

BLUE = "0F3D32"
GOLD = "C5A059"
WHITE = "FFFFFF"
LIGHT = "F4F7F5"
GREY = "D0D7D3"
GREEN = "C6EFCE"
YELLOW = "FFF2CC"
ORANGE = "FCE4D6"

HEADER_FILL = PatternFill("solid", fgColor=BLUE)
HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color=BLUE, size=16)
SUB_FONT = Font(name="Calibri", bold=True, color=BLUE, size=12)
BODY = Font(name="Calibri", size=10)
THIN = Side(style="thin", color=GREY)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def phase_of(d: date) -> tuple[str, str]:
    if date(2026, 4, 8) <= d <= date(2026, 5, 23):
        return "observe", "Observe · CRM · Sites · Sales dashboards"
    if date(2026, 5, 24) <= d <= date(2026, 6, 4):
        return "hdfc", "HDFC script with Vishal (delivered)"
    if date(2026, 6, 5) <= d <= date(2026, 6, 14):
        return "gift", "Flipkart gift cards with Vishal"
    if date(2026, 6, 15) <= d <= date(2026, 7, 1):
        return "auto-hire", "Flipkart automation + team hiring"
    return "hr-suite", "HR assigned + Flipkart suite / platform"


PHASE_TASKS = {
    "observe": [
        ("Observation & Onboarding", "FlightPath operations observation", "Shadowed workflows, tools, and handoffs."),
        ("CRM Solution", "CRM solution design & configuration", "Lead capture, stages, sales follow-ups."),
        ("Website Design", "Static website design (batch)", "Built toward ~20 static sites for FlightPath."),
        ("Sales Sheets & Dashboards", "Sales team sheets & dashboards", "Pipeline, targets, daily sales tracking."),
    ],
    "hdfc": [
        ("HDFC Script", "HDFC script with Vishal", "Development, hardening, delivery with Vishal."),
        ("HDFC Script", "HDFC debugging & dry-run", "Failure paths, logging, retry behaviour."),
        ("HDFC Script", "HDFC delivery & handoff", "Runbook and acceptance closed."),
        ("Team Support", "Pair review / support", "Environment and integration support."),
    ],
    "gift": [
        ("Flipkart Gift Cards", "Gift-card flow with Vishal", "Create/checkout gift-card paths."),
        ("Flipkart Gift Cards", "Gift-card UI & validation", "Balances, payment option wiring."),
        ("Flipkart Gift Cards", "Gift-card regression", "Regression and ops handoff notes."),
        ("Flipkart Automation", "Suite touchpoints", "Aligned gift-card module with suite contracts."),
    ],
    "auto-hire": [
        ("Flipkart Automation", "Automation suite build", "Checkout, tracking, accounts, queues, workspace."),
        ("Flipkart Automation", "Automation test & fixes", "Smoke paths, OTP/queue issues."),
        ("Team Hiring", "Team hiring support", "Screening, interviews, role fit."),
        ("Team Hiring", "Shortlist & feedback", "Written feedback for leadership/HR."),
    ],
    "hr-suite": [
        ("HR Assigned", "HR-assigned task execution", "Process, docs, compliance, coordination."),
        ("Flipkart Automation", "Flipkart suite delivery", "Checkout, tracking, invoices, OTP, agent, UI."),
        ("Platform & VPS", "VPS / platform ops", "Health, nginx/SSL, deploys, containers."),
        ("Internal Tools", "Internal tools / ID card", "FlightPath Infogain utilities."),
    ],
}

SLOTS = [
    ("09:00", "11:00", 2.0, 0),
    ("11:00", "13:00", 2.0, 1),
    ("13:30", "15:30", 2.0, 2),
    ("15:30", "17:30", 2.0, 3),
]


def workdays(start: date, end: date) -> list[date]:
    out: list[date] = []
    cur = start
    while cur <= end:
        if cur.weekday() != 6:  # exclude Sunday
            out.append(cur)
        cur += timedelta(days=1)
    return out


def build_rows(start: date, end: date) -> list[dict]:
    rows: list[dict] = []
    for d in workdays(start, end):
        phase_id, phase_label = phase_of(d)
        weekday = d.strftime("%A")
        month = d.strftime("%Y-%m")

        if d in LEAVE:
            rows.append(
                {
                    "date": d.isoformat(),
                    "weekday": weekday,
                    "month": month,
                    "status": "Leave",
                    "phase": phase_label,
                    "category": "—",
                    "start": "—",
                    "end": "—",
                    "hours": 0.0,
                    "overtime": "No",
                    "title": "Leave",
                    "detail": "Approved leave — no work hours.",
                }
            )
            continue

        if d in ON_MOVEMENT:
            rows.append(
                {
                    "date": d.isoformat(),
                    "weekday": weekday,
                    "month": month,
                    "status": "On Movement",
                    "phase": phase_label,
                    "category": "Planning & Review",
                    "start": "09:00",
                    "end": "17:30",
                    "hours": 0.0,
                    "overtime": "No",
                    "title": "On movement",
                    "detail": "Travel / movement — attendance only.",
                }
            )
            continue

        bank = PHASE_TASKS[phase_id]
        # Special day: 20 Jul — Tarun Sir Sector 8
        if d == date(2026, 7, 20):
            day_slots = [
                ("09:00", "11:00", 2.0, bank[0]),
                ("11:00", "13:00", 2.0, bank[1]),
                ("13:30", "15:30", 2.0, bank[2]),
                ("15:30", "17:30", 2.0, ("Planning & Review", "Prep for Tarun Sir meeting (Sector 8)", "Prepared discussion points for evening meeting.")),
                ("18:00", "19:30", 1.5, ("HR Assigned", "Meeting with Tarun Sir — Sector 8", "On-site meeting ~6:00 PM.", True)),
            ]
            for start_t, end_t, hrs, meta in day_slots:
                ot = False
                if len(meta) == 4:
                    cat, title, detail, ot = meta
                else:
                    cat, title, detail = meta
                rows.append(
                    {
                        "date": d.isoformat(),
                        "weekday": weekday,
                        "month": month,
                        "status": "Work",
                        "phase": phase_label,
                        "category": cat,
                        "start": start_t,
                        "end": end_t,
                        "hours": hrs,
                        "overtime": "Yes" if ot else "No",
                        "title": title,
                        "detail": detail,
                    }
                )
            continue

        # Special day: 21 Jul — MD Tarun + HR documentation
        if d == date(2026, 7, 21):
            day_slots = [
                (
                    "09:00",
                    "11:00",
                    2.0,
                    (
                        "Internal Tools",
                        "FlightPath daily work tracker & Excel export",
                        "Updated daily work log, Excel dashboard export (date/month), tracker repo.",
                    ),
                ),
                (
                    "11:00",
                    "13:00",
                    2.0,
                    (
                        "Internal Tools",
                        "ID card generator & digest automation setup",
                        "FlightPath Infogain ID-card tooling and daily digest automation (Gmail draft).",
                    ),
                ),
                (
                    "13:30",
                    "14:00",
                    0.5,
                    (
                        "Planning & Review",
                        "Prep for MD Tarun meeting & HR doc session",
                        "Prepared notes for MD meeting and HR documentation agenda.",
                    ),
                ),
                (
                    "14:00",
                    "14:30",
                    0.5,
                    ("HR Assigned", "Meeting with MD Tarun", "Meeting with MD Tarun (14:00–14:30)."),
                ),
                (
                    "14:30",
                    "17:30",
                    3.0,
                    (
                        "HR Assigned",
                        "HR documentation — account opening, visiting card, ID cards",
                        "Worked with HR on documentation for account opening, visiting cards, and ID cards.",
                    ),
                ),
            ]
            for start_t, end_t, hrs, meta in day_slots:
                cat, title, detail = meta
                rows.append(
                    {
                        "date": d.isoformat(),
                        "weekday": weekday,
                        "month": month,
                        "status": "Work",
                        "phase": phase_label,
                        "category": cat,
                        "start": start_t,
                        "end": end_t,
                        "hours": hrs,
                        "overtime": "No",
                        "title": title,
                        "detail": detail,
                    }
                )
            continue

        # Special day: 22 Jul — Pathwaynexgen mail, Sec 14 project, Sec 8 HR
        if d == date(2026, 7, 22):
            day_slots = [
                (
                    "09:00",
                    "11:00",
                    2.0,
                    (
                        "HR Assigned",
                        "Pathwaynexgen — add alternative mail ID",
                        "Configured / added alternative mail ID for the Pathwaynexgen mail account.",
                        False,
                    ),
                ),
                (
                    "11:00",
                    "13:00",
                    2.0,
                    (
                        "Internal Tools",
                        "Daily work log & agent collaboration",
                        "FlightPath daily tracker updates, Excel export, tooling with Cursor agent.",
                        False,
                    ),
                ),
                (
                    "13:30",
                    "15:30",
                    2.0,
                    (
                        "Flipkart Automation",
                        "Project delivery work",
                        "Continued FlightPath project work before evening site moves.",
                        False,
                    ),
                ),
                (
                    "15:30",
                    "17:30",
                    2.0,
                    (
                        "Flipkart Automation",
                        "Project work — wrap before Sector 14",
                        "Project progress ahead of 17:30 return to Sector 14.",
                        False,
                    ),
                ),
                (
                    "17:30",
                    "18:20",
                    0.8,
                    (
                        "Flipkart Automation",
                        "Sector 14 — project work",
                        "Returned to Sector 14 at 17:30; worked on the project.",
                        True,
                    ),
                ),
                (
                    "18:20",
                    "19:30",
                    1.2,
                    (
                        "HR Assigned",
                        "Sector 8 — helping HR",
                        "Back to Sector 8 from 18:20; supported HR until 19:30 (7:30 PM).",
                        True,
                    ),
                ),
            ]
            for start_t, end_t, hrs, meta in day_slots:
                cat, title, detail, ot = meta
                rows.append(
                    {
                        "date": d.isoformat(),
                        "weekday": weekday,
                        "month": month,
                        "status": "Work",
                        "phase": phase_label,
                        "category": cat,
                        "start": start_t,
                        "end": end_t,
                        "hours": hrs,
                        "overtime": "Yes" if ot else "No",
                        "title": title,
                        "detail": detail,
                    }
                )
            continue

        # Special day: 23 Jul — team support, reports, laptop asset, agent work (10–6, done 22:00)
        if d == date(2026, 7, 23):
            day_slots = [
                (
                    "10:00",
                    "12:00",
                    2.0,
                    (
                        "Internal Tools",
                        "Agent collaboration — FlightPath tracker & tooling",
                        "Worked with Cursor agent on daily work log, Excel/tracker updates, and tooling.",
                        False,
                    ),
                ),
                (
                    "12:00",
                    "14:00",
                    2.0,
                    (
                        "Team Support",
                        "Helping team & assigning tasks",
                        "Supported team members and assigned tasks with owners and expectations.",
                        False,
                    ),
                ),
                (
                    "14:30",
                    "16:30",
                    2.0,
                    (
                        "Team Support",
                        "Collect reports & analyse",
                        "Took team reports, reviewed status, and analysed findings for follow-ups.",
                        False,
                    ),
                ),
                (
                    "16:30",
                    "18:00",
                    1.5,
                    (
                        "HR Assigned",
                        "Laptop asset purchase support",
                        "Helped with laptop / IT asset purchase process (specs, quotes, or approvals).",
                        False,
                    ),
                ),
                (
                    "18:00",
                    "22:00",
                    4.0,
                    (
                        "Internal Tools",
                        "Complete daily task update with agent",
                        "Closed today’s work log / tracker update with Cursor agent; complete at ~22:00.",
                        True,
                    ),
                ),
            ]
            for start_t, end_t, hrs, meta in day_slots:
                cat, title, detail, ot = meta
                rows.append(
                    {
                        "date": d.isoformat(),
                        "weekday": weekday,
                        "month": month,
                        "status": "Work",
                        "phase": phase_label,
                        "category": cat,
                        "start": start_t,
                        "end": end_t,
                        "hours": hrs,
                        "overtime": "Yes" if ot else "No",
                        "title": title,
                        "detail": detail,
                    }
                )
            continue

        for start_t, end_t, hrs, idx in SLOTS:
            cat, title, detail = bank[idx % len(bank)]
            rows.append(
                {
                    "date": d.isoformat(),
                    "weekday": weekday,
                    "month": month,
                    "status": "Work",
                    "phase": phase_label,
                    "category": cat,
                    "start": start_t,
                    "end": end_t,
                    "hours": hrs,
                    "overtime": "No",
                    "title": title,
                    "detail": detail,
                }
            )

        # Beyond hours on crunch phases / Fri-Sat
        heavy = phase_id in ("hdfc", "gift") or (
            phase_id in ("auto-hire", "hr-suite", "observe") and d.weekday() >= 4
        )
        if heavy:
            ot_map = {
                "hdfc": ("HDFC Script", "HDFC after-hours delivery push", "Continued with Vishal past roster."),
                "gift": ("Flipkart Gift Cards", "Gift-card after-hours pairing", "Extended session past 17:30."),
                "auto-hire": ("Flipkart Automation", "Automation after-hours build/test", "Extra hours past roster."),
                "observe": ("Website Design", "Sites/dashboards weekend catch-up", "Beyond roster catch-up."),
                "hr-suite": ("Flipkart Automation", "Suite/HR after-hours wrap-up", "Beyond roster wrap-up."),
            }
            cat, title, detail = ot_map[phase_id]
            rows.append(
                {
                    "date": d.isoformat(),
                    "weekday": weekday,
                    "month": month,
                    "status": "Work",
                    "phase": phase_label,
                    "category": cat,
                    "start": "18:00",
                    "end": "20:00",
                    "hours": 2.0,
                    "overtime": "Yes",
                    "title": title,
                    "detail": detail,
                }
            )

    return rows


def newest_first(rows: list[dict]) -> list[dict]:
    """Latest dates first; within a day keep chronological start time."""
    by_time = sorted(rows, key=lambda r: r["start"] if r["start"] != "—" else "00:00")
    return sorted(by_time, key=lambda r: r["date"], reverse=True)


def style_header(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 10
        for cell in col:
            if cell.value is None:
                continue
            width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def write_dashboard(wb: Workbook, rows: list[dict], label: str) -> None:
    ws = wb.create_sheet("01 Dashboard", 0)
    ws["A1"] = "FlightPath Infogain — Daily Work Dashboard"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = f"Filter: {label} · Excludes Sundays, ProductiveIT, personal · 8h + 30m lunch"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")
    ws.merge_cells("A2:F2")

    work_rows = [r for r in rows if r["status"] == "Work"]
    leave_days = len({r["date"] for r in rows if r["status"] == "Leave"})
    move_days = len({r["date"] for r in rows if r["status"] == "On Movement"})
    work_days = len({r["date"] for r in work_rows})
    roster_h = sum(r["hours"] for r in work_rows if r["overtime"] == "No")
    ot_h = sum(r["hours"] for r in work_rows if r["overtime"] == "Yes")

    metrics = [
        ("Working days", work_days),
        ("Roster hours", roster_h),
        ("Beyond hours (OT)", ot_h),
        ("Leave days", leave_days),
        ("On movement days", move_days),
        ("Total logged hours", roster_h + ot_h),
    ]
    ws["A4"] = "KPI"
    ws["B4"] = "Value"
    style_header(ws, 4, 2)
    for i, (k, v) in enumerate(metrics, start=5):
        ws.cell(i, 1, k).font = BODY
        ws.cell(i, 1).border = BORDER
        ws.cell(i, 2, v).font = BODY
        ws.cell(i, 2).border = BORDER
        ws.cell(i, 2).alignment = CENTER
        if i % 2 == 0:
            ws.cell(i, 1).fill = PatternFill("solid", fgColor=LIGHT)
            ws.cell(i, 2).fill = PatternFill("solid", fgColor=LIGHT)

    # Month summary table for chart
    ws["D4"] = "Month"
    ws["E4"] = "Roster hours"
    ws["F4"] = "OT hours"
    style_header(ws, 4, 6)
    months = sorted({r["month"] for r in rows})
    for i, m in enumerate(months, start=5):
        subset = [r for r in work_rows if r["month"] == m]
        ws.cell(i, 4, m).border = BORDER
        ws.cell(i, 5, sum(r["hours"] for r in subset if r["overtime"] == "No")).border = BORDER
        ws.cell(i, 6, sum(r["hours"] for r in subset if r["overtime"] == "Yes")).border = BORDER

    if months:
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.title = "Hours by month (roster vs OT)"
        chart.y_axis.title = "Hours"
        chart.x_axis.title = "Month"
        data = Reference(ws, min_col=5, min_row=4, max_col=6, max_row=4 + len(months))
        cats = Reference(ws, min_col=4, min_row=5, max_row=4 + len(months))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart.height = 10
        chart.width = 15
        ws.add_chart(chart, "A12")

    # Category hours for pie
    cat_start = 5 + max(len(months), 6) + 2
    ws.cell(cat_start, 4, "Category").font = SUB_FONT
    ws.cell(cat_start + 1, 4, "Category")
    ws.cell(cat_start + 1, 5, "Hours")
    style_header(ws, cat_start + 1, 5)
    from collections import Counter

    cat_hours: Counter[str] = Counter()
    for r in work_rows:
        if r["category"] != "—":
            cat_hours[r["category"]] += r["hours"]
    for i, (cat, hrs) in enumerate(sorted(cat_hours.items(), key=lambda x: -x[1]), start=cat_start + 2):
        ws.cell(i, 4, cat).border = BORDER
        ws.cell(i, 5, hrs).border = BORDER

    if cat_hours:
        pie = PieChart()
        pie.title = "Hours by work stream"
        last = cat_start + 1 + len(cat_hours)
        labels = Reference(ws, min_col=4, min_row=cat_start + 2, max_row=last)
        data = Reference(ws, min_col=5, min_row=cat_start + 1, max_row=last)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.height = 10
        pie.width = 12
        ws.add_chart(pie, "D12")

    # Phase summary
    ws["H4"] = "Phase"
    ws["I4"] = "Work days"
    ws["J4"] = "Roster h"
    ws["K4"] = "OT h"
    style_header(ws, 4, 11)
    phases = []
    for r in rows:
        if r["phase"] not in phases:
            phases.append(r["phase"])
    for i, ph in enumerate(phases, start=5):
        days = {r["date"] for r in work_rows if r["phase"] == ph}
        subset = [r for r in work_rows if r["phase"] == ph]
        ws.cell(i, 8, ph).border = BORDER
        ws.cell(i, 9, len(days)).border = BORDER
        ws.cell(i, 10, sum(r["hours"] for r in subset if r["overtime"] == "No")).border = BORDER
        ws.cell(i, 11, sum(r["hours"] for r in subset if r["overtime"] == "Yes")).border = BORDER

    ws["A28"] = "How to filter in Excel"
    ws["A28"].font = SUB_FONT
    ws["A29"] = "1) Open sheet '02 Daily Log' — use Excel AutoFilter on Date / Month / Status / Category."
    ws["A30"] = "2) Open sheet '03 Monthly Summary' for month-wise totals."
    ws["A31"] = "3) Re-export with: python export_excel.py --month 2026-07   or   --date 2026-07-20"
    ws["A32"] = "4) ProductiveIT and personal work are excluded by design."
    autosize(ws)


def write_daily(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("02 Daily Log")
    headers = [
        "Date",
        "Weekday",
        "Month",
        "Status",
        "Phase",
        "Category",
        "Start",
        "End",
        "Hours",
        "Beyond hours?",
        "Task",
        "Detail",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header(ws, 1, len(headers))

    for r_i, r in enumerate(rows, start=2):
        vals = [
            r["date"],
            r["weekday"],
            r["month"],
            r["status"],
            r["phase"],
            r["category"],
            r["start"],
            r["end"],
            r["hours"],
            r["overtime"],
            r["title"],
            r["detail"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r_i, c, v)
            cell.font = BODY
            cell.border = BORDER
            cell.alignment = CENTER if c <= 10 else LEFT
        if r["status"] == "Leave":
            fill = PatternFill("solid", fgColor=YELLOW)
            for c in range(1, 13):
                ws.cell(r_i, c).fill = fill
        elif r["status"] == "On Movement":
            fill = PatternFill("solid", fgColor=ORANGE)
            for c in range(1, 13):
                ws.cell(r_i, c).fill = fill
        elif r["overtime"] == "Yes":
            ws.cell(r_i, 10).fill = PatternFill("solid", fgColor=GREEN)

    last = 1 + len(rows)
    if last > 1:
        table = Table(displayName="DailyLog", ref=f"A1:L{last}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True
        )
        ws.add_table(table)
    ws.auto_filter.ref = f"A1:L{last}"
    ws.freeze_panes = "A2"
    autosize(ws)


def write_monthly(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("03 Monthly Summary")
    headers = [
        "Month",
        "Work days",
        "Leave days",
        "Movement days",
        "Roster hours",
        "OT hours",
        "Total hours",
        "Lunch breaks (min)",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header(ws, 1, len(headers))

    months = sorted({r["month"] for r in rows})
    for i, m in enumerate(months, start=2):
        subset = [r for r in rows if r["month"] == m]
        work_days = {r["date"] for r in subset if r["status"] == "Work"}
        leave_days = {r["date"] for r in subset if r["status"] == "Leave"}
        move_days = {r["date"] for r in subset if r["status"] == "On Movement"}
        work_rows = [r for r in subset if r["status"] == "Work"]
        roster = sum(r["hours"] for r in work_rows if r["overtime"] == "No")
        ot = sum(r["hours"] for r in work_rows if r["overtime"] == "Yes")
        vals = [m, len(work_days), len(leave_days), len(move_days), roster, ot, roster + ot, len(work_days) * 30]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i, c, v)
            cell.font = BODY
            cell.border = BORDER
            cell.alignment = CENTER
    ws.freeze_panes = "A2"
    autosize(ws)


def write_by_date(wb: Workbook, rows: list[dict]) -> None:
    """One compact day rollup for quick date filter / pivot."""
    ws = wb.create_sheet("04 Day Rollup")
    headers = ["Date", "Weekday", "Month", "Status", "Phase", "Roster hours", "OT hours", "Task count", "Focus / notes"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header(ws, 1, len(headers))

    dates = []
    for r in rows:
        if r["date"] not in dates:
            dates.append(r["date"])
    # rows already newest-first; keep day rollup in that order

    for i, d in enumerate(dates, start=2):
        subset = [r for r in rows if r["date"] == d]
        status = subset[0]["status"]
        phase = subset[0]["phase"]
        weekday = subset[0]["weekday"]
        month = subset[0]["month"]
        if status != "Work":
            focus = subset[0]["title"]
            roster = ot = 0
            count = 0
        else:
            roster = sum(r["hours"] for r in subset if r["overtime"] == "No")
            ot = sum(r["hours"] for r in subset if r["overtime"] == "Yes")
            count = len(subset)
            focus = next((r["title"] for r in subset if r["overtime"] == "Yes"), subset[-1]["title"])
        vals = [d, weekday, month, status, phase, roster, ot, count, focus]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i, c, v)
            cell.font = BODY
            cell.border = BORDER
            cell.alignment = CENTER if c < 9 else LEFT
        if status == "Leave":
            for c in range(1, 10):
                ws.cell(i, c).fill = PatternFill("solid", fgColor=YELLOW)
        elif status == "On Movement":
            for c in range(1, 10):
                ws.cell(i, c).fill = PatternFill("solid", fgColor=ORANGE)

    last = 1 + len(dates)
    if last > 1:
        table = Table(displayName="DayRollup", ref=f"A1:I{last}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(table)
    ws.auto_filter.ref = f"A1:I{last}"
    ws.freeze_panes = "A2"
    autosize(ws)


def write_howto(wb: Workbook) -> None:
    ws = wb.create_sheet("00 How to use")
    ws["A1"] = "Export & filter guide"
    ws["A1"].font = TITLE_FONT
    lines = [
        "",
        "Sheets",
        "• 01 Dashboard — KPIs, month chart, category pie, phase table",
        "• 02 Daily Log — every task slot (filter by Date / Month / Status / Category)",
        "• 03 Monthly Summary — month-wise totals",
        "• 04 Day Rollup — one row per day (best for date-wise review)",
        "",
        "Date-wise in Excel",
        "1. Open 04 Day Rollup or 02 Daily Log",
        "2. Click the filter arrow on the Date column",
        "3. Pick a single date or a range",
        "",
        "Month-wise in Excel",
        "1. Filter the Month column (YYYY-MM) on Daily Log or Day Rollup",
        "2. Or open 03 Monthly Summary",
        "",
        "Re-export a focused file from terminal",
        "  python export_excel.py",
        "  python export_excel.py --month 2026-07",
        "  python export_excel.py --date 2026-07-20",
        "  python export_excel.py --from 2026-05-24 --to 2026-06-14",
        "",
        "Windows: double-click EXPORT-EXCEL.bat",
        "",
        "Excluded by design: ProductiveIT*, personal/*, Sundays",
        "Recipient review email: virrdhiman@gmail.com (DRAFT only until you finalize)",
    ]
    for i, line in enumerate(lines, start=2):
        ws.cell(i, 1, line).font = BODY
    ws.column_dimensions["A"].width = 100


def export(start: date, end: date, label: str, outfile: Path) -> Path:
    rows = newest_first(build_rows(start, end))
    wb = Workbook()
    # remove default
    default = wb.active
    wb.remove(default)
    write_howto(wb)
    write_dashboard(wb, rows, label)
    write_daily(wb, rows)
    write_monthly(wb, rows)
    write_by_date(wb, rows)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    outfile.parent.mkdir(parents=True, exist_ok=True)
    wb.save(outfile)
    return outfile


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Export FlightPath daily work log to Excel")
    p.add_argument("--month", help="YYYY-MM (e.g. 2026-07)")
    p.add_argument("--date", help="Single day YYYY-MM-DD")
    p.add_argument("--from", dest="date_from", help="Range start YYYY-MM-DD")
    p.add_argument("--to", dest="date_to", help="Range end YYYY-MM-DD")
    p.add_argument("-o", "--output", help="Output .xlsx path")
    return p.parse_args()


def main() -> None:
    args = parse_args()
    start, end = START, END
    label = f"{START.isoformat()} to {END.isoformat()}"
    stamp = f"{START.isoformat()}_to_{END.isoformat()}"

    if args.date:
        d = date.fromisoformat(args.date)
        start = end = d
        label = f"Date {d.isoformat()}"
        stamp = d.isoformat()
    elif args.month:
        y, m = map(int, args.month.split("-"))
        start = date(y, m, 1)
        if m == 12:
            end = date(y, 12, 31)
        else:
            end = date(y, m + 1, 1) - timedelta(days=1)
        start = max(start, START)
        end = min(end, END)
        label = f"Month {args.month}"
        stamp = args.month
    elif args.date_from or args.date_to:
        start = date.fromisoformat(args.date_from) if args.date_from else START
        end = date.fromisoformat(args.date_to) if args.date_to else END
        label = f"{start.isoformat()} to {end.isoformat()}"
        stamp = f"{start.isoformat()}_to_{end.isoformat()}"

    outfile = Path(args.output) if args.output else OUT_DIR / f"FlightPath_Daily_Work_{stamp}.xlsx"
    path = export(start, end, label, outfile)
    print(f"Wrote {path}")
    print(f"Filter: {label}")
    print("Open in Excel -> use AutoFilter on Daily Log / Day Rollup, or Dashboard charts.")


if __name__ == "__main__":
    main()
